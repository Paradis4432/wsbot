from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import openpyxl
from pathlib import Path
import traceback
import re
import base64
import requests
from PIL import Image
from selenium.common.exceptions import NoSuchElementException
from paramiko import SSHClient
from scp import SCPClient
from dotenv import load_dotenv
import os
import json

folder = "negocioCholo"


load_dotenv()

ssh = SSHClient()
ssh.load_system_host_keys()
ssh.connect(os.getenv('SERVERIP'), username='root',
            password=os.getenv('SSHPASS'))

# SCPCLient takes a paramiko transport as an argument
scp = SCPClient(ssh.get_transport())

driver = webdriver.Chrome()
driver.get("https://web.whatsapp.com")

time.sleep(10)
print("starting")
cont = 1

p = Path(f"./backup/{folder}.xlsx")
dataExcel = openpyxl.load_workbook(p)
sheet = dataExcel.active

with open("cont.txt", "r") as f:
    cont = int(f.read())
# cont = 0
with open("subcont.txt", "r") as f:
    subcont = int(f.read())
# subcont = 0
print(f"found {cont} {subcont}")
newLine = False


def hasStar():
    star_el = driver.find_element(
        By.XPATH, '//*[@id="app"]/div/span[3]/div/div/div[2]/div/div[1]/div[2]/div/div[2]/div/span')

    return star_el.get_attribute("data-testid") == "star-btn"


def clickNextButton():
    try:
        driver.find_element(
            By.XPATH, '//*[@id="app"]/div/span[3]/div/div/div[2]/div/div[2]/div[1]/div/div').click()
    except Exception as e:
        print("finished")
        driver.find_element(
            By.XPATH, '//*[@id="app"]/div/span[3]/div/div/div[2]/div/div[1]/div[2]/div/div[7]/div/span').click()


keys = ['ShYSmkGi19ztn5G7oDBPLUbN', 'gcLYS11Z6jrL7MQP65mZ2y7C', 'ds4c2TkgKKNZwJ1qSJ4sCtAG',
        '5awd6A82Jthc1jA383R9z3VP', 'zMPPPMiD2GCPVMDmfDdNDeLh', 'eoSRfgSTAN75o6AQ1YUQ87h4']
currentKey = 0

def has0left():
    global currentKey

    print("3")

    headers = {
        'accept': '*/*',
        'X-API-Key': keys[currentKey],
    }

    response = requests.get(
        'https://api.remove.bg/v1.0/account', headers=headers)
    callsLeft = json.loads(response.text)[
        "data"]["attributes"]["api"]["free_calls"]
    print("4")

    # print(response.text, keys[currentKey])
    print(f"found {keys[currentKey]} with {callsLeft} left")

    print("5")

    return callsLeft == 0


def cleanIMG():
    global currentKey

    print("1")
    while has0left():
        print("2")

        if currentKey == len(keys) - 1:
            currentKey = 0
        else:
            currentKey += 1

    try:
        print("6")

        print("using key: " + keys[currentKey])
        response = requests.post(
            'https://api.remove.bg/v1.0/removebg',
            files={'image_file': open(f"backup/{folder}/img{cont}.{subcont}.png", 'rb')},
            data={'size': 'auto'},
            headers={'X-Api-Key': keys[currentKey]}
        )

        if response.status_code == requests.codes.ok:
            with open(f"./backup/{folder}/img{cont}.{subcont}.png", 'wb') as out:
                out.write(response.content)
        else:
            print("Error:", response.status_code, response.text)
            time.sleep(3)

        try:
            img = Image.open(f"backup/{folder}/img{cont}.{subcont}.png").convert("RGBA")
            x, y = img.size
            card = Image.new("RGBA", (x, y), (255, 255, 255))
            card.paste(img, (0, 0, x, y), img)
            card.save(f"backup/{folder}/img{cont}.{subcont}.png", format="png")
        except Exception as e:
            print("error pasting white bg")
            print(e)

    except Exception as e:
        print("error cleaning img:")
        print(e)



    """ try:
        # no need to place white bg with bg_color: 'white' in request
        img = Image.open("images/img" + str(cont) + "." +
                         str(subcont) + ".png").convert("RGBA")
        x, y = img.size
        card = Image.new("RGBA", (x, y), (255, 255, 255))
        card.paste(img, (0, 0, x, y), img)
        card.save("images/img" + str(cont) + "." +
                  str(subcont) + ".png", format="png")
    except Exception as e:
        print("error cleaning img:")
        print(e) """


def saveInExcel(datos, ip):
    c = cont + 1
    sheet.cell(row=c, column=1).value = datos[0]
    sheet.cell(row=c, column=2).value = datos[1]
    sheet.cell(row=c, column=3).value = datos[2]

    x = 4
    for i in ips:
        sheet.cell(row=c, column=x).value = i
        sheet.cell(row=c, column=x).hyperlink = (i)
        x += 1

    dataExcel.save(f"./backup/{folder}.xlsx")



ips = []
datos = []
i,o,e = ssh.exec_command(f"mkdir /root/wsbot/backup/{folder}")
if e.read():
    print(e.read())
    time.sleep(5)

while True:
    try:
        if hasStar():
            time.sleep(1)
            clickNextButton()
            continue

        # save cont and subcont
        print(f"saving {cont} {subcont}")
        with open("cont.txt", "w") as f:
            f.write('%d' % cont)

        with open("subcont.txt", "w") as f:
            f.write('%d' % subcont)

        result = driver.execute_async_script("""
                        var uri = arguments[0];
                        var callback = arguments[1];
                        var toBase64 = function(buffer){for(var r,n=new Uint8Array(buffer),t=n.length,a=new Uint8Array(4*Math.ceil(t/3)),i=new Uint8Array(64),o=0,c=0;64>c;++c)i[c]="ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789+/".charCodeAt(c);for(c=0;t-t%3>c;c+=3,o+=4)r=n[c]<<16|n[c+1]<<8|n[c+2],a[o]=i[r>>18],a[o+1]=i[r>>12&63],a[o+2]=i[r>>6&63],a[o+3]=i[63&r];return t%3===1?(r=n[t-1],a[o]=i[r>>2],a[o+1]=i[r<<4&63],a[o+2]=61,a[o+3]=61):t%3===2&&(r=(n[t-2]<<8)+n[t-1],a[o]=i[r>>10],a[o+1]=i[r>>4&63],a[o+2]=i[r<<2&63],a[o+3]=61),new TextDecoder("ascii").decode(a)};
                        var xhr = new XMLHttpRequest();
                        xhr.responseType = 'arraybuffer';
                        xhr.onload = function(){ callback(toBase64(xhr.response)) };
                        xhr.onerror = function(){ callback(xhr.status) };
                        xhr.open('GET', uri);
                        xhr.send();
                        """, driver.find_element(By.XPATH, '//*[@id="app"]/div/span[3]/div/div/div[2]/div/div[2]/div[2]/div/div/div/div/div[2]/div/img').get_attribute("src"))

        try:
            img_text_el = driver.find_element(
                By.XPATH, '//*[@id="app"]/div/span[3]/div/div/div[2]/div/div[2]/div[2]/p/span')

        except NoSuchElementException:
            subcont += 1
        else:
            # tomar texto
            img_text = img_text_el.text
            datos = [int(s) for s in re.findall(r'\d+', img_text)]

            # new line
            newLine = True
            subcont += 1

        with open(f"backup/{folder}/img{cont}.{subcont}.png", 'wb') as f:
            f.write(base64.b64decode(result))

        cleanIMG()

        # removes star
        driver.find_element(
            By.XPATH, '//*[@id="app"]/div/span[3]/div/div/div[2]/div/div[1]/div[2]/div/div[2]/div/span').click()

        time.sleep(1)
        print("uploading")
        scp.put(f"backup/{folder}/img{cont}.{subcont}.png",
                f"/root/wsbot/backup/{folder}/img{cont}.{subcont}.png")
        ip = f"http://50.116.47.159/backup/{folder}/img{cont}.{subcont}.png"
        ips.append(ip)

        clickNextButton()

        if newLine:
            saveInExcel(datos, ips)
            ips = []

            subcont = 0
            cont += 1
            newLine = False

        time.sleep(1)

    except Exception as e:
        print("error")
        print(e)
        traceback.print_exc()
        time.sleep(1)
    finally:
        scp.close()
