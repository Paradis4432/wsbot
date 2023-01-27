import shutil
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import openpyxl
from pathlib import Path
import traceback
import re
import base64
import requests
from PIL import Image
from selenium.common.exceptions import NoSuchElementException
from paramiko import SSHClient
from scp import SCPClient
from dotenv import load_dotenv
import os

load_dotenv()

ssh = SSHClient()
ssh.load_system_host_keys()
ssh.connect(os.getenv('SERVERIP'), username='root',
            password=os.getenv('SSHPASS'))

# SCPCLient takes a paramiko transport as an argument
scp = SCPClient(ssh.get_transport())

print("changin")
options = webdriver.ChromeOptions()
prefs = {"download.default_directory": r"C:\Users\Lucas\Escritorio\e\wsbot\eraseBG\tmp\\"}
options.add_experimental_option("prefs", prefs)

driver = webdriver.Chrome(options=options)


driver.get("https://google.com")


driver.execute_script("window.open('https://web.whatsapp.com', 'ws');")

driver.execute_script("window.open('https://www.erase.bg/upload', 'erase');")

time.sleep(10)
# driver.switch_to.window("ws")
# time.sleep(3)
# driver.switch_to.window("erase")


print("starting")
cont = 1

p = Path("./valoresWsBot.xlsx")
dataExcel = openpyxl.load_workbook(p)
sheet = dataExcel.active

with open("cont.txt", "r") as f:
    cont = int(f.read())
# cont = 0
with open("subcont.txt", "r") as f:
    subcont = int(f.read())
# subcont = 0
print(f"found {cont} {subcont}")
newLine = False


def hasStar():
    star_el = driver.find_element(
        By.XPATH, '//*[@id="app"]/div/span[3]/div/div/div[2]/div/div[1]/div[2]/div/div[2]/div/span')

    return star_el.get_attribute("data-testid") == "star-btn"


def clickNextButton():
    try:
        driver.find_element(
            By.XPATH, '//*[@id="app"]/div/span[3]/div/div/div[2]/div/div[2]/div[1]/div/div').click()
    except Exception as e:
        print("finished")
        driver.find_element(
            By.XPATH, '//*[@id="app"]/div/span[3]/div/div/div[2]/div/div[1]/div[2]/div/div[7]/div/span').click()


def cleanIMG():

    driver.switch_to.window("erase")

    driver.find_element(By.XPATH, '//*[@id="PasteURL__HomePage"]').click()
    driver.find_element(By.XPATH, '//*[@id="modal-root"]/div[1]/div/div[1]/div[1]/input').send_keys(
        f"http://50.116.47.159/images/img{cont}.{subcont}.png")
    driver.find_element(
        By.XPATH, '//*[@id="modal-root"]/div[1]/div/div[1]/div[1]/button').click()

    attemps = 0
    time.sleep(2)
    try:
        # if ad is found
        driver.find_element('//*[@id="modal-root"]/div[2]/div/div[2]').click()
    except NoSuchElementException:
        print()
    except Exception as e:
        print("error closing ad")
        print(e)

    while True:
        if attemps > 30:
            return
        try:
            # if X is found
            close = driver.find_element(
                By.XPATH, '//*[@id="root"]/div/div[1]/div[3]/div[2]/div/div[2]/div/div/div[3]')

        except NoSuchElementException as e:
            attemps += 1
            time.sleep(1)

    # download
    driver.find_element(
        By.XPATH, '//*[@id="root"]/div/div[1]/div[3]/div[2]/div/div[2]/div/div/div[2]/div[1]/div[1]/button').click()

    # mv and rename img
    os.rename(f'tmp/{os.listdir("./tmp")[0]}', f"tmp/img{cont}.{subcont}.png")
    shutil.move(f"tmp/img{cont}.{subcont}.png",
                f"images/img{cont}.{subcont}.png")
    # os.remove("tmp/img0.1.png")

    # close current img
    close = driver.find_element(
        By.XPATH, '//*[@id="root"]/div/div[1]/div[3]/div[2]/div/div[2]/div/div/div[3]').click()

    driver.switch_to.window("ws")

    # paste white img as bg
    try:
        img = Image.open(f"images/img{cont}.{subcont}.png").convert("RGBA")
        x, y = img.size
        card = Image.new("RGBA", (x, y), (255, 255, 255))
        card.paste(img, (0, 0, x, y), img)
        card.save(f"images/img{cont}.{subcont}.png", format="png")
        
    except Exception as e:
        print("error cleaning img:")
        print(e)


def saveInExcel(datos, ip):
    c = cont + 1
    sheet.cell(row=c, column=1).value = datos[0]
    sheet.cell(row=c, column=2).value = datos[1]
    sheet.cell(row=c, column=3).value = datos[2]

    x = 4
    for i in ips:
        sheet.cell(row=c, column=x).value = i
        sheet.cell(row=c, column=x).hyperlink = (i)
        x += 1

    dataExcel.save("valoresWsBot.xlsx")


ips = []
datos = []

while True:
    try:
        if hasStar():
            time.sleep(1)
            clickNextButton()
            continue

        # save cont and subcont
        print(f"saving {cont} {subcont}")
        with open("cont.txt", "w") as f:
            f.write('%d' % cont)

        with open("subcont.txt", "w") as f:
            f.write('%d' % subcont)

        result = driver.execute_async_script("""
                        var uri = arguments[0];
                        var callback = arguments[1];
                        var toBase64 = function(buffer){for(var r,n=new Uint8Array(buffer),t=n.length,a=new Uint8Array(4*Math.ceil(t/3)),i=new Uint8Array(64),o=0,c=0;64>c;++c)i[c]="ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789+/".charCodeAt(c);for(c=0;t-t%3>c;c+=3,o+=4)r=n[c]<<16|n[c+1]<<8|n[c+2],a[o]=i[r>>18],a[o+1]=i[r>>12&63],a[o+2]=i[r>>6&63],a[o+3]=i[63&r];return t%3===1?(r=n[t-1],a[o]=i[r>>2],a[o+1]=i[r<<4&63],a[o+2]=61,a[o+3]=61):t%3===2&&(r=(n[t-2]<<8)+n[t-1],a[o]=i[r>>10],a[o+1]=i[r>>4&63],a[o+2]=i[r<<2&63],a[o+3]=61),new TextDecoder("ascii").decode(a)};
                        var xhr = new XMLHttpRequest();
                        xhr.responseType = 'arraybuffer';
                        xhr.onload = function(){ callback(toBase64(xhr.response)) };
                        xhr.onerror = function(){ callback(xhr.status) };
                        xhr.open('GET', uri);
                        xhr.send();
                        """, driver.find_element(By.XPATH, '//*[@id="app"]/div/span[3]/div/div/div[2]/div/div[2]/div[2]/div/div/div/div/div[2]/div/img').get_attribute("src"))

        try:
            img_text_el = driver.find_element(
                By.XPATH, '//*[@id="app"]/div/span[3]/div/div/div[2]/div/div[2]/div[2]/p/span')

        except NoSuchElementException:
            subcont += 1
        else:
            # tomar texto
            img_text = img_text_el.text
            datos = [int(s) for s in re.findall(r'\d+', img_text)]

            # new line
            newLine = True
            subcont += 1

        with open(f"images/img{cont}.{subcont}.png"), 'wb' as f:
            f.write(base64.b64decode(result))

        print("uploading")
        scp.put(f"images/img{cont}.{subcont}.png",
                f"/root/wsbot/images/img{cont}.{subcont}.png")
        ip = f"http://50.116.47.159/images/img{cont}.{subcont}.png"
        ips.append(ip)

        cleanIMG()

        # removes star
        driver.find_element(
            By.XPATH, '//*[@id="app"]/div/span[3]/div/div/div[2]/div/div[1]/div[2]/div/div[2]/div/span').click()

        clickNextButton()

        if newLine:
            saveInExcel(datos, ips)
            ips = []

            subcont = 0
            cont += 1
            newLine = False

        time.sleep(1)

    except Exception as e:
        print("error")
        print(e)
        # traceback.print_exc()
        time.sleep(1)
    finally:
        scp.close()
